import transformacoes
import pandas as pd
from openpyxl import load_workbook
from scripts.path_names import path_planilha_banco_de_dados, path_planilha_principal, path_planilha_integrada

def gerar_planilha_integrada():
    workbook = load_workbook(path_planilha_banco_de_dados)
    banco_de_dados = workbook.active
    ultima_linha = banco_de_dados.max_row
    ultima_coleta = banco_de_dados[f"A{ultima_linha}"].value

    df_principal = transformacoes.tratar_planilha(pd.read_excel(path_planilha_principal), ultima_coleta)

    # Gravar as novas células na tabela do Banco de Dados
    for coluna in df_principal.columns:
        for index, celula in enumerate(df_principal[f"{coluna}"], start=ultima_linha+1):
            banco_de_dados[f"{coluna}{index}"] = celula

    for row in banco_de_dados.iter_rows(min_row=ultima_linha+1):
        banco_de_dados.row_dimensions[row[0].row].height = 60

    workbook.save(path_planilha_integrada)

if __name__ == '__main__':
    gerar_planilha_integrada()
