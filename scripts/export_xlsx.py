import transformations
import pandas as pd
from openpyxl import load_workbook


workbook = load_workbook('../planilhas/banco_de_dados.xlsx')
banco_de_dados = workbook.active
ultima_linha = banco_de_dados.max_row
ultima_coleta = banco_de_dados[f"A{ultima_linha}"].value+1
df_principal = transformations.tratar_planilha(pd.read_excel('../planilhas/planilha_principal.xlsx'), ultima_coleta)

# Gravar as novas células na tabela do Banco de Dados

for coluna in df_principal.columns:
    for index, celula in enumerate(df_principal[f"{coluna}"], start=ultima_linha+1):
        banco_de_dados[f"{coluna}{index}"] = celula

for row in banco_de_dados.iter_rows(min_row=ultima_linha+1):
    banco_de_dados.row_dimensions[row[0].row].height = 60

workbook.save('../planilhas/teste.xlsx')
